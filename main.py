"""
main.py
-------
End-to-end HR Attrition Analysis Pipeline
==========================================
Run with:  python main.py

Steps:
  1. Generate / load IBM HR Analytics dataset
  2. Data preprocessing (clean, encode, split, scale)
  3. Exploratory Data Analysis (6 visualisation sets)
  4. Model training: Logistic Regression + Gradient Boosting + Random Forest
  5. Model evaluation: accuracy, ROC-AUC, confusion matrix, PR curves
  6. Feature importance + SHAP-proxy analysis
  7. 6-Month attrition trend forecasting
  8. Export results summary to Excel
"""
import os
import sys
import time
import warnings

warnings.filterwarnings("ignore")

# ── path setup ──────────────────────────────────────────────
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

from src.generate_data       import generate_hr_dataset
from src.data_preprocessing  import clean, encode, split_and_scale
from src.eda                 import run_eda
from src.model               import run_modeling
from src.forecasting         import run_forecasting


def export_summary(results: list[dict], history, forecast):
    """Write model performance + forecast summary to Excel."""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Sheet 1: Model Performance ──────────────────────
        ws1 = wb.active
        ws1.title = "Model Performance"
        header_fill = PatternFill("solid", fgColor="1565C0")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["Model", "Accuracy", "ROC-AUC", "Avg Precision"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, res in enumerate(results, 2):
            ws1.cell(row=row, column=1, value=res["name"])
            ws1.cell(row=row, column=2, value=round(res["acc"], 4))
            ws1.cell(row=row, column=3, value=round(res["auc"], 4))
            ws1.cell(row=row, column=4, value=round(res["ap"],  4))

        for col in range(1, 5):
            ws1.column_dimensions[get_column_letter(col)].width = 30

        # ── Sheet 2: Forecast ────────────────────────────────
        ws2 = wb.create_sheet("6-Month Forecast")
        ws2.append(["Month", "Forecasted Attrition Rate (%)"])
        ws2["A1"].font = Font(bold=True)
        ws2["B1"].font = Font(bold=True)
        for _, row in forecast.iterrows():
            ws2.append([row["Month"].strftime("%b %Y"), round(row["AttritionRate"] * 100, 2)])
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 28

        path = os.path.join("outputs", "summary_report.xlsx")
        wb.save(path)
        print(f"  ✓  Summary exported → {path}")
    except Exception as e:
        print(f"  ⚠  Excel export skipped: {e}")


def main():
    t0 = time.time()
    os.makedirs("data",    exist_ok=True)
    os.makedirs("outputs", exist_ok=True)

    print("=" * 58)
    print("  HR Analytics & Employee Attrition Analysis Pipeline")
    print("=" * 58)

    # ── 1. Data ──────────────────────────────────────────────
    print("\n[1/5] Generating IBM HR Analytics dataset…")
    df_raw = generate_hr_dataset()
    df_raw.to_csv("data/WA_Fn-UseC_-HR-Employee-Attrition.csv", index=False)
    print(f"      {df_raw.shape[0]} employees · {df_raw.shape[1]} features")
    print(f"      Attrition rate: {(df_raw['Attrition']=='Yes').mean():.1%}")

    # ── 2. Preprocessing ────────────────────────────────────
    print("\n[2/5] Preprocessing…")
    df_clean   = clean(df_raw)
    df_enc, _  = encode(df_clean)
    X_tr, X_te, y_tr, y_te, X_tr_sc, X_te_sc, scaler = split_and_scale(df_enc)
    print(f"      Train {X_tr.shape} | Test {X_te.shape}")

    # ── 3. EDA ───────────────────────────────────────────────
    print("\n[3/5] Running EDA…")
    run_eda(df_raw)

    # ── 4. Modeling ──────────────────────────────────────────
    print("\n[4/5] Training models…")
    lr, gbm, rf, all_results = run_modeling(X_tr, X_te, y_tr, y_te, X_tr_sc, X_te_sc)

    # ── 5. Forecasting ───────────────────────────────────────
    print("\n[5/5] Forecasting…")
    history, forecast = run_forecasting(df_raw, gbm, X_te, y_te)

    # ── Export ───────────────────────────────────────────────
    export_summary(all_results, history, forecast)

    elapsed = time.time() - t0
    print(f"\n{'='*58}")
    print(f"  ✅  Pipeline complete in {elapsed:.1f}s")
    print(f"  📁  All outputs saved to outputs/")
    best = max(all_results, key=lambda r: r["acc"])
    print(f"  🏆  Best model: {best['name']}")
    print(f"      Accuracy={best['acc']:.4f}  |  ROC-AUC={best['auc']:.4f}")
    print("=" * 58)


if __name__ == "__main__":
    main()
